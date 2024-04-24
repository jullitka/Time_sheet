import calendar
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border,
                             Font, PatternFill, Side)
from openpyxl.worksheet.pagebreak import Break

from add_data import search_employee, search_all_employees
from utils import (check_holiday, duration,
                   generate_name,
                   start_or_end_filling, reporting_period)


wb = Workbook()
ws = wb.active

# ориентация листа
ws.page_setup.orientation = 'landscape'
# размер листа
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.row_breaks.append(Break(id=49))
ws.row_breaks.append(Break(id=97))

row_number = 42
page_break_1 = Break(id=row_number)
ws.row_breaks.append(page_break_1)
row_number = 82
page_break_2 = Break(id=row_number)
ws.row_breaks.append(page_break_2)

# название столбцов для дат
LETTERS_FOR_DATE = ('E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                    'N', 'O', 'P', 'Q', 'R', 'S', 'T')

# название столбцов шириной 4
LETTERS_FOR_4 = ('X', 'Y', 'AA', 'AB', 'AC', 'AE')

# название столбцов шириной 5
LETTERS_FOR_5 = ('W', 'Z', 'AD', 'AF')

# название столбцов шириной 6.5
LETTERS_FOR_6_5 = ('D', 'V')

# название столбцов для нумерации
LETTERS_FOR_NUM = ('B', 'C', 'D', 'E', 'U', 'V', 'W', 'X', 'Y',
                   'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF')

# определим стили сторон
# тонкая черная линия
stile_thin = Side(border_style="thin", color="000000")
# толстая черная линия
stile_medium = Side(border_style="medium", color="000000")

sum_days = ''
sum_hours = ''
sum_ill = ''
sum_not_salary = ''


def row_dimension(num, time_sheet_name):
    """Устанавливает высоту строк"""
    for i in range(num+9):
        ws.row_dimensions[i].height = 11
    wb.save(time_sheet_name)


def document_name(time_sheet_name):
    """Создает надпись названия документа"""
    ws.merge_cells('H5:L5')
    ws['H5'].value = 'ТАБЕЛЬ'
    ws['H5'].font = Font(name='Arial', bold=True, size=10)
    ws['H5'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('E6:O6')
    ws['E6'].value = 'УЧЕТА РАБОЧЕГО ВРЕМЕНИ'
    ws['E6'].font = Font(name='Arial', bold=True, size=10)
    ws['E6'].alignment = Alignment(horizontal="center", vertical="center")
    wb.save(time_sheet_name)


def name_company(department_name, time_sheet_name):
    """Создает надпись с названием"""
    ws.merge_cells('C1:Z2')
    ws['C1'].value = ('ФГУП "Всероссийский научно-исследовательский '
                      'институт физико-технических и радиотехнических '
                      'измерений"\n'
                      f'НИО-7 {department_name}')
    ws['C1'].font = Font(name='Arial', bold=True, underline='single', size=9)
    ws['C1'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    wb.save(time_sheet_name)


def table_num_date(time_sheet_name):
    """Создает таблицу с датой заполнения"""
    ws.merge_cells('Q5:T6')
    ws['Q5'].value = 'номер документа'
    ws['Q5'].font = Font(name='Calibri', size=7)
    ws['Q5'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('U5:V6')
    ws['U5'].value = 'дата заполнения'
    ws['U5'].font = Font(name='Calibri', size=7)
    ws['U5'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('Q7:T7')
    ws.merge_cells('U7:V7')
    ws['U7'].fill = PatternFill('solid', fgColor="E6B8B7")
    for row in range(5, 8):
        for col in range(17, 23):
            ws.cell(row=row, column=col).border = Border(
                            top=stile_thin,
                            bottom=stile_thin,
                            left=stile_thin,
                            right=stile_thin
            )
    wb.save(time_sheet_name)


def table_report_period(time_sheet_name):
    """Создает таблицу "отчетный период"."""
    ws.merge_cells('X5:AA5')
    ws['X5'].value = 'отчётный период'
    ws['X5'].font = Font(name='Calibri', size=7)
    ws['X5'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('X6:Y6')
    ws['X6'].value = 'с'
    ws['X6'].font = Font(name='Calibri', size=7)
    ws['X6'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('Z6:AA6')
    ws['Z6'].value = 'по'
    ws['Z6'].font = Font(name='Calibri', size=7)
    ws['Z6'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('Z7:AA7')
    ws.merge_cells('X7:Y7')
    for col in range(24, 28):
        for row in range(5, 8):
            ws.cell(row=row, column=col).border = Border(
                            top=stile_thin,
                            bottom=stile_thin,
                            left=stile_thin,
                            right=stile_thin
            )
    ws['X7'].fill = PatternFill('solid', fgColor="FFFF00")
    ws['Z7'].fill = PatternFill('solid', fgColor="E6B8B7")
    wb.save(time_sheet_name)


def form_info(time_sheet_name):
    """Создает запись с информацией о форме
    в правом верхнем углу"""
    # первая строчка
    ws.merge_cells('AD2:AF2')
    ws['AD2'].value = 'Форма № К-13'
    ws['AD2'].font = Font(name='Calibri', size=7)
    ws['AD2'].alignment = Alignment(horizontal="right", vertical="center")
    # вторая строка
    ws.merge_cells('Z3:AF3')
    ws['Z3'].value = 'Утверждена приказом генерального директора'
    ws['Z3'].font = Font(name='Calibri', size=7)
    ws['Z3'].alignment = Alignment(horizontal="right", vertical="center")
    # третья строка
    ws.merge_cells('Z4:AF4')
    ws['Z4'].value = 'ФГУП "ВНИИФТРИ" от 29 декабря 2012 г. №247'
    ws['Z4'].font = Font(name='Calibri', size=7)
    ws['Z4'].alignment = Alignment(horizontal="right", vertical="center")
    wb.save(time_sheet_name)


def table_header(time_sheet_name):
    """Создает шапку таблицы без расставления дат"""
    # объединение ячеек по вертикали
    ws.merge_cells('B10:B16')
    ws.merge_cells('C10:C16')
    ws.merge_cells('D10:D16')
    ws.merge_cells('U11:U13')
    ws.merge_cells('V11:V13')
    ws.merge_cells('W15:W16')
    ws.merge_cells('X15:X16')
    ws.merge_cells('Y15:Y16')
    ws.merge_cells('Z15:Z16')
    ws.merge_cells('AA15:AA16')
    ws.merge_cells('AB15:AB16')
    ws.merge_cells('AC10:AF11')
    ws.merge_cells('AC12:AC16')
    ws.merge_cells('AD12:AD16')
    ws.merge_cells('AE12:AE16')
    ws.merge_cells('AF12:AF16')

    # объединение ячеек для колонок с датами
    for let in LETTERS_FOR_DATE:
        string_1 = f'{let}11:{let}13'
        string_2 = f'{let}14:{let}16'
        ws.merge_cells(string_1)
        ws.merge_cells(string_2)

    # объединение ячеек по горизонтали
    ws.merge_cells('E10:T10')
    ws.merge_cells('E17:T17')
    ws.merge_cells('U10:V10')
    ws.merge_cells('U14:V16')
    ws.merge_cells('W10:AB10')
    ws.merge_cells('W11:AB11')
    ws.merge_cells('W12:AB12')
    ws.merge_cells('W13:AB13')
    ws.merge_cells('W14:Y14')
    ws.merge_cells('Z14:AB14')

    # установка ширины столбцов
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["U"].width = 5.7
    for col in LETTERS_FOR_DATE:
        ws.column_dimensions[col].width = 2.8
    for let in LETTERS_FOR_4:
        ws.column_dimensions[let].width = 4
    for let in LETTERS_FOR_5:
        ws.column_dimensions[let].width = 5
    for let in LETTERS_FOR_6_5:
        ws.column_dimensions[let].width = 6.5

    # записываем текст в ячейки
    ws['B10'].value = '№ п/п'
    ws['B10'].font = Font(name='Arial', size=8)
    ws['B10'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['C10'].value = 'Фамилия, инициалы, должность'
    ws['C10'].font = Font(name='Arial', bold=True, size=9)
    ws['C10'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['D10'].value = 'Таб. №'
    ws['D10'].font = Font(name='Arial', bold=True, size=10)
    ws['D10'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['E10'].value = 'Отметки о явках и неявках на работу по числам месяца'
    ws['E10'].font = Font(name='Arial', bold=True, size=8)
    ws['E10'].alignment = Alignment(horizontal="center", vertical="center")

    ws['U10'].value = 'Отработано за'
    ws['U10'].font = Font(name='Arial', bold=True, size=8)
    ws['U10'].alignment = Alignment(horizontal="center", vertical="center")

    ws['U11'].value = 'половину мес.'
    ws['U11'].font = Font(name='Arial', size=6)
    ws['U11'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['V11'].value = 'месяц'
    ws['V11'].alignment = Alignment(horizontal="center", vertical="center")
    ws['V11'].font = Font(name='Arial', size=7)

    ws['U14'].value = 'дни/часы'
    ws['U14'].font = Font(name='Arial', size=6)
    ws['U14'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['W10'].value = 'Данные для начисл. з/п по видам и напр. затрат'
    ws['W10'].font = Font(name='Arial', size=5.5)
    ws['W10'].alignment = Alignment(horizontal="center", vertical="center")

    ws['W11'].value = 'код вида оплат'
    ws['W11'].font = Font(name='Arial', size=7)
    ws['W11'].alignment = Alignment(horizontal="center", vertical="center")

    ws['W13'].value = 'корреспондирующий счет'
    ws['W13'].font = Font(name='Arial', size=7)
    ws['W13'].alignment = Alignment(horizontal="center", vertical="center")

    ws['W15'].value = 'код вида оплат'
    ws['W15'].font = Font(name='Arial', size=6)
    ws['W15'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['X15'].value = 'коpp. счет'
    ws['X15'].font = Font(name='Arial', size=6)
    ws['X15'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['Y15'].value = 'дни (часы)'
    ws['Y15'].font = Font(name='Arial', size=6)
    ws['Y15'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['Z15'].value = 'код вида оплат'
    ws['Z15'].font = Font(name='Arial', size=6)
    ws['Z15'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['AA15'].value = 'коpp. счет'
    ws['AA15'].font = Font(name='Arial', size=6)
    ws['AA15'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['AB15'].value = 'дни (часы)'
    ws['AB15'].font = Font(name='Arial', size=6)
    ws['AB15'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['AC10'].value = 'Неявки по причинам'
    ws['AC10'].font = Font(name='Arial', bold=True, size=7)
    ws['AC10'].alignment = Alignment(horizontal="center", vertical="center")

    ws['AC12'].value = 'код'
    ws['AC12'].font = Font(name='Arial', size=6)
    ws['AC12'].alignment = Alignment(horizontal="center", vertical="center")

    ws['AD12'].value = 'дни (часы)'
    ws['AD12'].font = Font(name='Arial', size=6)
    ws['AD12'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    ws['AE12'].value = 'код'
    ws['AE12'].font = Font(name='Arial', size=6)
    ws['AE12'].alignment = Alignment(horizontal="center", vertical="center")

    ws['AF12'].value = 'дни (часы)'
    ws['AF12'].font = Font(name='Arial', size=6)
    ws['AF12'].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )
    for i in range(len(LETTERS_FOR_NUM)):
        string = f'{LETTERS_FOR_NUM[i]}17'
        ws[string].font = Font(name='Arial', size=6, italic=True)
        ws[string].value = i + 1
        ws[string].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrapText=True
        )
    # рисование границ шапки таблицы
    for row in range(10, 18):
        for col in range(2, 33):
            if row == 10 and col == 2:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                bottom=stile_thin,
                                                left=stile_medium,
                                                right=stile_thin
                )
            elif row == 17 and col == 2:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_medium,
                                                left=stile_medium,
                                                right=stile_thin
                )
            elif row == 10 and col in (4, 20, 22, 28, 32):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_medium
                )
            elif row == 17 and col in (4, 20, 22, 28, 32):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_medium,
                                                left=stile_thin,
                                                right=stile_medium
                )
            elif row == 10:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_thin
                )
            elif row == 17:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_medium,
                                                left=stile_thin,
                                                right=stile_thin
                )
            elif col == 2:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_thin,
                                                left=stile_medium,
                                                right=stile_thin
                )
            elif col in (4, 20, 22, 28, 32):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_medium
                )
            else:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_thin
                )
    wb.save(time_sheet_name)


def table_to_one_emploee(num, count, time_sheet_name):
    """Создает таблицу для одного сотрудника"""
    global sum_days
    global sum_hours
    global sum_ill
    global sum_not_salary
    ws.merge_cells(f'B{num}:B{num+3}')
    ws[f'B{num}'].value = f'{count}'
    ws[f'B{num}'].font = Font(name='Arial', size=8)
    ws[f'B{num}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f'C{num}:C{num+1}')
    # ws.merge_cells(f'C{num+2}:C{num+3}')
    ws.merge_cells(f'D{num}:D{num+3}')
    ws.merge_cells(f'V{num}:V{num+1}')
    ws.merge_cells(f'V{num+2}:V{num+3}')
    ws.merge_cells(f'Y{num}:Y{num+1}')
    ws.merge_cells(f'Y{num+2}:Y{num+3}')
    ws.merge_cells(f'AB{num}:AB{num+1}')
    ws.merge_cells(f'AB{num+2}:AB{num+3}')

    # заливка цветом
    for row in range(4):
        for cell in (f'AD{num+row}', f'AF{num+row}'):
            ws[cell].fill = PatternFill('solid', fgColor="99FFFF")
    ws[f'V{num}'].fill = PatternFill('solid', fgColor="99FFFF")
    ws[f'V{num+2}'].fill = PatternFill('solid', fgColor="99FFFF")
    # ячейки для больничного, формула
    ws[f'AC{num}'].value = 'Б'
    ws[f'AC{num}'].font = Font(name='Arial', size=8)
    ws[f'AC{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'AD{num}'] = (f"=COUNTIF(E{num}:T{num},AC{num})"
                      f"+COUNTIF(E{num+2}:T{num+2},AC{num})")
    ws[f'AD{num}'].font = Font(name='Arial', size=8)
    ws[f'AD{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    sum_ill += f'+AD{num}'

    # командировки
    ws[f'AE{num}'].value = 'К'
    ws[f'AE{num}'].font = Font(name='Arial', size=8)
    ws[f'AE{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'AF{num}'] = (f"=COUNTIF(E{num}:T{num},AE{num})"
                      f"+COUNTIF(E{num+2}:T{num+2},AE{num})")
    ws[f'AF{num}'].font = Font(name='Arial', size=8)
    ws[f'AF{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # отпуск
    ws[f'AC{num+2}'].value = 'ОТ'
    ws[f'AC{num+2}'].font = Font(name='Arial', size=8)
    ws[f'AC{num+2}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'AD{num+2}'] = (f"=COUNTIF(E{num}:T{num},AC{num+2})"
                        f"+COUNTIF(E{num+2}:T{num+2},AC{num+2})")
    ws[f'AD{num+2}'].font = Font(name='Arial', size=8)
    ws[f'AD{num+2}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # без сохранения з/п
    ws[f'AE{num+2}'].value = 'ДО'
    ws[f'AE{num+2}'].font = Font(name='Arial', size=8)
    ws[f'AE{num+2}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'AF{num+2}'] = (f"=COUNTIF(E{num}:T{num},AE{num+2})"
                        f"+COUNTIF(E{num+2}:T{num+2},AE{num+2})")
    ws[f'AF{num+2}'].font = Font(name='Arial', size=8)
    ws[f'AF{num+2}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    sum_not_salary += f'+AF{num+2}'
    # рассчет количества отработанных дней
    # за первую половину месяца
    ws[f'U{num}'] = f'=COUNTIF(E{num}:T{num},"Я")+COUNTIF(E{num}:T{num},"ОН")'
    ws[f'U{num}'].font = Font(name='Arial', size=8)
    ws[f'U{num}'].alignment = Alignment(horizontal="center", vertical="center")
    # рассчет количества отработанных часов
    # за первую половину месяца
    ws[f'U{num+1}'] = f'=SUM(E{num+1}:T{num+1})'
    ws[f'U{num+1}'].font = Font(name='Arial', size=8)
    ws[f'U{num+1}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # рассчет количества отработанных дней
    # за вторую половину месяца
    ws[f'U{num+2}'] = (f'=COUNTIF(E{num+2}:T{num+2},"Я")'
                       f'+COUNTIF(E{num+2}:T{num+2},"ОН")')
    ws[f'U{num+2}'].font = Font(name='Arial', size=8)
    ws[f'U{num+2}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # рассчет количества отработанных часов
    # за вторую половину месяца
    ws[f'U{num+3}'] = f'=SUM(E{num+3}:T{num+3})'
    ws[f'U{num+3}'].font = Font(name='Arial', size=8)
    ws[f'U{num+3}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # рассчет количества отработанных дней за месяц
    ws[f'V{num}'] = f'=U{num}+U{num+2}'
    ws[f'V{num}'].font = Font(name='Arial', size=8)
    ws[f'V{num}'].alignment = Alignment(horizontal="center", vertical="center")
    sum_days += f'+V{num}'
    # рассчет количества отработанных часов за месяца
    ws[f'V{num+2}'] = f'=U{num+1}+U{num+3}'
    ws[f'V{num+2}'].font = Font(name='Arial', size=8)
    ws[f'V{num+2}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    sum_hours += f'+V{num+2}'

    for row in range(num, num+4):
        for col in range(2, 33):
            if col == 3 and row == num:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                left=stile_medium,
                                                right=stile_medium
                )
            elif col == 3 and row in (num+1, num+2):
                ws.cell(row=row, column=col).border = Border(
                                                left=stile_medium,
                                                right=stile_medium
                )
            elif col == 3 and row == num+3:
                ws.cell(row=row, column=col).border = Border(
                                                bottom=stile_medium,
                                                left=stile_medium,
                                                right=stile_medium
                )
            elif row in (num, num+2) and col in (2, 4, 5, 21, 22, 29):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                bottom=stile_thin,
                                                left=stile_medium,
                                                right=stile_thin
                )
            elif row == num+1 and col in (2, 4, 5, 21, 22, 29):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_thin,
                                                left=stile_medium,
                                                right=stile_thin
                )
            elif row == num+3 and col in (2, 4, 5, 21, 22, 29):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_medium,
                                                left=stile_medium,
                                                right=stile_thin
                )
            elif row in (num, num+2) and col == 32:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_medium
                )
            elif row == num+1 and col == 32:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_medium
                )
            elif row == num+3 and col == 32:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_medium,
                                                left=stile_thin,
                                                right=stile_medium
                )
            elif row in (num, num+2):
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_medium,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_thin
                )
            elif row == num+3:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_medium,
                                                left=stile_thin,
                                                right=stile_thin
                )
            else:
                ws.cell(row=row, column=col).border = Border(
                                                top=stile_thin,
                                                bottom=stile_thin,
                                                left=stile_thin,
                                                right=stile_thin
                )
    wb.save(time_sheet_name)


def total_sums(num, time_sheet_name):
    """Итоговые суммы по дням, часам, больничным и т.д."""
    # заливка цветом нужных ячеек
    ws[f'V{num}'].fill = PatternFill('solid', fgColor="99FFFF")
    ws[f'V{num+1}'].fill = PatternFill('solid', fgColor="99FFFF")
    ws[f'AD{num}'].fill = PatternFill('solid', fgColor="99FFFF")
    ws[f'AD{num+1}'].fill = PatternFill('solid', fgColor="99FFFF")

    ws.merge_cells(f'S{num+4}:U{num+4}')
    ws[f'S{num}'].value = 'ИТОГО:'
    ws[f'S{num}'].font = Font(name='Arial', bold=True, size=8)
    ws[f'S{num}'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f'X{num}:AB{num}')
    ws[f'X{num}'].value = 'ИТОГО: больничных '
    ws[f'X{num}'].font = Font(name='Arial', bold=True, size=8)
    ws[f'X{num}'].alignment = Alignment(horizontal="right", vertical="center")

    ws[f'AC{num}'].value = '(Б)'
    ws[f'AC{num}'].font = Font(name='Arial', bold=True, size=8)
    ws[f'AC{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws.merge_cells(f'Y{num+1}:AB{num+1}')
    ws[f'Y{num+1}'].value = 'без сохранения З/П '
    ws[f'Y{num+1}'].font = Font(name='Arial', size=8)
    ws[f'Y{num+1}'].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    ws[f'AC{num+1}'].value = '(ДО)'
    ws[f'AC{num+1}'].font = Font(name='Arial', bold=True, size=8)
    ws[f'AC{num+1}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'W{num}'].value = 'дней'
    ws[f'W{num}'].font = Font(name='Arial', size=8)
    ws[f'W{num}'].alignment = Alignment(horizontal="center", vertical="center")

    ws[f'W{num+1}'].value = 'часов'
    ws[f'W{num+1}'].font = Font(name='Arial', size=8)
    ws[f'W{num+1}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'AE{num}'].value = 'дней'
    ws[f'AE{num}'].font = Font(name='Arial', size=8)
    ws[f'AE{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'AE{num+1}'].value = 'дней'
    ws[f'AE{num+1}'].font = Font(name='Arial', size=8)
    ws[f'AE{num+1}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws[f'V{num}'].border = Border(bottom=stile_thin)
    ws[f'AD{num}'].border = Border(bottom=stile_thin)
    # итоговый рассчет количества отработанных дней
    ws[f'V{num}'] = f'=SUM({sum_days})'
    ws[f'V{num}'].font = Font(name='Arial', size=8)
    ws[f'V{num}'].alignment = Alignment(horizontal="center", vertical="center")
    # итоговый рассчет количества отработанных часов
    ws[f'V{num+1}'] = f'=SUM({sum_hours})'
    ws[f'V{num+1}'].font = Font(name='Arial', size=8)
    ws[f'V{num+1}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # итоговый рассчет дней больничного
    ws[f'AD{num}'] = f'=SUM({sum_ill})'
    ws[f'AD{num}'].font = Font(name='Arial', size=8)
    ws[f'AD{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    # итоговый рассчет без сохранения з/п
    ws[f'AD{num+1}'] = f'=SUM({sum_not_salary})'
    ws[f'AD{num+1}'].font = Font(name='Arial', size=8)
    ws[f'AD{num+1}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    wb.save(time_sheet_name)


def signature(num, position, keeps_time_sheet,
              head_of_department, head_of_HR, time_sheet_name):
    """Оформляет места для полписей"""
    ws.merge_cells(f'D{num+3}:E{num+3}')
    ws.merge_cells(f'D{num+4}:E{num+4}')
    ws.merge_cells(f'F{num+4}:I{num+4}')
    ws.merge_cells(f'J{num+3}:N{num+3}')
    ws.merge_cells(f'J{num+4}:N{num+4}')
    ws.merge_cells(f'R{num+3}:U{num+3}')
    ws.merge_cells(f'R{num+4}:U{num+4}')
    ws.merge_cells(f'R{num+6}:U{num+6}')
    ws.merge_cells(f'R{num+7}:U{num+7}')
    ws.merge_cells(f'V{num+4}:W{num+4}')
    ws.merge_cells(f'V{num+7}:W{num+7}')
    ws.merge_cells(f'X{num+3}:AA{num+3}')
    ws.merge_cells(f'X{num+4}:AA{num+4}')
    ws.merge_cells(f'X{num+6}:AA{num+6}')
    ws.merge_cells(f'X{num+7}:AA{num+7}')

    ws[f'C{num+3}'].value = 'Ответственное лицо '
    ws[f'C{num+3}'].font = Font(name='Arial', size=8)
    ws[f'C{num+3}'].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    # должность ответственного лица
    ws[f'D{num+3}'].value = position
    ws[f'D{num+3}'].font = Font(name='Arial', size=8)
    ws[f'D{num+3}'].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )
    for cell in (f'D{num+4}', f'R{num+4}', f'R{num+7}'):
        ws[cell].value = 'должность'
        ws[cell].font = Font(name='Arial', size=6)
        ws[cell].alignment = Alignment(
            horizontal="left",
            vertical="center"
        )
    for cell in (f'F{num+4}', f'V{num+4}', f'V{num+7}'):
        ws[cell].value = 'подпись'
        ws[cell].font = Font(name='Arial', size=6)
        ws[cell].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
    for cell in (f'J{num+4}', f'X{num+4}', f'X{num+7}'):
        ws[cell].value = 'расшифровка подписи'
        ws[cell].font = Font(name='Arial', size=6)
        ws[cell].alignment = Alignment(
            horizontal="right",
            vertical="center"
        )
    # фамилия ответственного лица
    ws[f'J{num+3}'].value = keeps_time_sheet
    ws[f'J{num+3}'].font = Font(name='Arial', size=8)
    ws[f'J{num+3}'].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    # фамилия начальника отдела
    ws[f'X{num+3}'].value = head_of_department
    ws[f'X{num+3}'].font = Font(name='Arial', size=8)
    ws[f'X{num+3}'].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    # фамилия начальника отдела кадров
    ws[f'X{num+6}'].value = head_of_HR
    ws[f'X{num+6}'].font = Font(name='Arial', size=8)
    ws[f'X{num+6}'].alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    ws[f'R{num+3}'].value = 'Начальник отдела'
    ws[f'R{num+3}'].font = Font(name='Arial', size=8)
    ws[f'R{num+3}'].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )
    ws[f'R{num+6}'].value = 'Нач. отд. кадров'
    ws[f'R{num+6}'].font = Font(name='Arial', size=8)
    ws[f'R{num+6}'].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )
    for col in range(4, 15):
        ws.cell(row=num+3, column=col).border = Border(bottom=stile_thin)
    for col in range(18, 28):
        ws.cell(row=num+3, column=col).border = Border(bottom=stile_thin)
    for col in range(18, 28):
        ws.cell(row=num+6, column=col).border = Border(bottom=stile_thin)
    wb.save(time_sheet_name)


def signature_date(num, time_sheet_name):
    """Проставляет даты подписи"""
    ws.merge_cells(f'AD{num+3}:AE{num+3}')
    ws.merge_cells(f'AD{num+6}:AE{num+6}')
    for cell in (f'AD{num+3}', f'AD{num+6}'):
        ws[cell] = '=U7'
        ws[cell].fill = PatternFill('solid', fgColor="99FFFF")
        ws[cell].font = Font(name='Arial', size=8)
        ws[cell].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
    wb.save(time_sheet_name)


def fill_dates(year, month, time_sheet_name):
    """Заполняет шапку датами и закрашивает
    выходные и праздничные дни"""
    num_days = calendar.monthrange(year, month)

    for row in (11, 14):
        if row == 11:
            for col in range(5, 21):
                if col - 4 <= 15:
                    ws.cell(row=row, column=col).value = col - 4
                    if check_holiday(year, month, col - 4):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            'solid', fgColor="66CCFF"
                        )
                else:
                    ws.cell(row=row, column=col).value = 'X'
                ws.cell(row=row, column=col).font = Font(name='Arial', size=8)
                ws.cell(row=row, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                )
        elif row == 14:
            for col in range(5, 21):
                if col + 11 <= num_days[1]:
                    ws.cell(row=row, column=col).value = col + 11
                    if check_holiday(year, month, col + 11):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            'solid',
                            fgColor="66CCFF"
                        )
                else:
                    ws.cell(row=row, column=col).value = 'X'
                ws.cell(row=row, column=col).font = Font(name='Arial', size=8)
                ws.cell(row=row, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                )
    wb.save(time_sheet_name)


def add_data_to_time_sheet(employee: tuple, num: int, file_name_absence,
                           year, month, full_month, time_sheet_name):
    """Добавляет данные о сотруднике в таблицу табеля"""
    last_name = employee[1]
    first_name = employee[2]
    patronymic = employee[3]
    time_sheet_number = employee[4]
    position = employee[5]
    part_time = employee[6]
    internal_combine = employee[7]
    employment_date = employee[8]
    dismissal_date = employee[9]
    # department = employee[8]
    absence_employees = search_all_employees(file_name_absence, year, month)
    name = generate_name(last_name, first_name, patronymic)

    if internal_combine == '*' or internal_combine == '-нет-':
        # если не внутренний совместитель, объединим ячейки
        ws.merge_cells(f'C{num+2}:C{num+3}')
        # заполняет ячейку с именем сотрудника
        ws[f'C{num}'].value = name
        ws[f'C{num}'].font = Font(name='Arial', size=9)
        ws[f'C{num}'].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        # добавляет должность
        ws[f'C{num+2}'].value = position
        ws[f'C{num+2}'].font = Font(name='Arial', size=8)
        ws[f'C{num+2}'].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
    elif internal_combine == 'совм':
        # если внутренний совместитель
        # заполняет ячейку с именем сотрудника
        ws[f'C{num}'].value = name
        ws[f'C{num}'].font = Font(name='Arial', size=9)
        ws[f'C{num}'].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        # добавляет должность
        ws[f'C{num+2}'].value = position
        ws[f'C{num+2}'].font = Font(name='Arial', size=8)
        ws[f'C{num+2}'].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        # добавляет информацию о внутреннем совместительстве
        ws[f'C{num+3}'].value = '(0,5 ст. внутр. совм.)'
        ws[f'C{num+3}'].font = Font(name='Arial', size=8)
        ws[f'C{num+3}'].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
    # заполняет ячейку с табельным номером
    ws[f'D{num}'].value = time_sheet_number
    ws[f'D{num}'].font = Font(name='Arial', size=8)
    ws[f'D{num}'].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    absence = search_employee(absence_employees, name, internal_combine)
    # перебираем столбцы, в которые нужно записывать явки
    for col in range(5, 21):
        # число месяца, обрабатываемое в данных ячейках
        # для первой половины месяца
        day_1 = col - 4
        # для второй половины месяца
        day_2 = col + 11
        # даты пропуска рабочих дней
        

        if check_holiday(year, month, day_1):
            # и закрашиваем соответствующие ячейки
            ws.cell(row=num, column=col).fill = PatternFill(
                        'solid', fgColor="66CCFF"
            )
            ws.cell(row=num+1, column=col).fill = PatternFill(
                        'solid', fgColor="66CCFF"
            )
        if check_holiday(year, month, day_2):
            # и закрашиваем соответствующие ячейки
            ws.cell(row=num+2, column=col).fill = PatternFill(
                      'solid', fgColor="66CCFF"
            )
            ws.cell(row=num+3, column=col).fill = PatternFill(
                       'solid', fgColor="66CCFF"
                )
        # числа первой половины месяца меньше 15
        # но больше дня троусдтойства, если он был в этом же месяце
        if (start_or_end_filling(
            year, month, employment_date
        ).day <= day_1 <= 15 and day_1 <= start_or_end_filling(
            year, month, dismissal_date, dismissal=True
        ).day):
            # если в этот день сотрудник отсутствовал
            if datetime(year, month, day_1) in absence:
                abs_date = datetime(year, month, day_1)
                # записываем в ячейку код отсутствия
                ws.cell(row=num, column=col).value = absence[abs_date]
                ws.cell(row=num, column=col).font = Font(name='Arial', size=8)
                ws.cell(row=num, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                )
            # если это выходной день, записываем в ячейку 'В'
            elif check_holiday(year, month, day_1):
                ws.cell(row=num, column=col).value = 'В'
                ws.cell(row=num, column=col).font = Font(name='Arial', size=8)
                ws.cell(row=num, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                )
            # в остальных случаях записываем 'Я'
            else:
                ws.cell(row=num, column=col).value = 'Я'
                ws.cell(row=num, column=col).font = Font(name='Arial', size=8)
                ws.cell(row=num, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                )
                ws.cell(row=num+1, column=col).value = duration(
                    year, month, day_1, part_time
                )
                ws.cell(row=num+1, column=col).font = Font(
                    name='Arial', size=8
                )
                ws.cell(row=num+1, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                )
        # последний столбец всегда заполнен 'X'
        else:
            if (start_or_end_filling(
                year, month, employment_date
            ).day <= day_1 <= start_or_end_filling(
                year, month, dismissal_date, dismissal=True
            ).day):
                for i in (num, num+1):
                    ws.cell(i, column=col).value = 'X'
                    ws.cell(i, column=col).font = Font(name='Arial', size=8)
                    ws.cell(i, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                    )
    # если нужно заполнять весь месяц, заполняем еще и вторую половину
    if full_month:
        # перебираем столбцы, в которые нужно записывать явки
        for col in range(5, 21):
            # для второй половины месяца
            day_2 = col + 11
            # если даты из интервала трудоустройства и увольнения
            if (start_or_end_filling(
                year, month, employment_date
            ).day <= day_2 <= start_or_end_filling(
                year, month, dismissal_date, dismissal=True
            ).day):
                # если в этот день сотрудник отсутствовал
                if datetime(year, month, day_2) in absence:
                    abs_date = datetime(year, month, day_2)
                    # записываем в ячейку код отсутствия
                    ws.cell(row=num+2, column=col).value = absence[abs_date]
                    ws.cell(row=num+2, column=col).font = Font(
                        name='Arial', size=8
                    )
                    ws.cell(row=num+2, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                    )
                # если это выходной день, записываем в ячейку 'В'
                elif check_holiday(year, month, day_2):
                    ws.cell(row=num+2, column=col).value = 'В'
                    ws.cell(row=num+2, column=col).font = Font(
                        name='Arial', size=8
                    )
                    ws.cell(row=num+2, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                    )
                # в остальных случаях записываем 'Я'
                else:
                    ws.cell(row=num+2, column=col).value = 'Я'
                    ws.cell(row=num+2, column=col).font = Font(
                        name='Arial', size=8
                    )
                    ws.cell(row=num+2, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                    )
                    ws.cell(row=num+3, column=col).value = duration(
                        year, month, day_2, part_time
                    )
                    ws.cell(row=num+3, column=col).font = Font(
                        name='Arial', size=8
                    )
                    ws.cell(row=num+3, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                    )
            # если еще не все ячейки заполнены, а месяц кончился
            else:
                # то заполняем их 'X'
                for i in (num+2, num+3):
                    ws.cell(i, column=col).value = 'X'
                    ws.cell(i, column=col).font = Font(name='Arial', size=8)
                    ws.cell(i, column=col).alignment = Alignment(
                                    horizontal="center",
                                    vertical="center"
                    )
    wb.save(time_sheet_name)


def put_down_dates(year, month, report_date, time_sheet_name):
    """Вносит дату заполнения табеля
    и даты отчетного периода"""
    reporting_period(year, month, report_date)
    ws['U7'].value = report_date
    ws['U7'].font = Font(name='Calibri', size=7)
    ws['U7'].alignment = Alignment(horizontal="center", vertical="center")

    ws['X7'].value = datetime(year, month, 1).strftime('%d.%m.%Y')
    ws['X7'].font = Font(name='Calibri', size=7)
    ws['X7'].alignment = Alignment(horizontal="center", vertical="center")

    ws['Z7'].value = reporting_period(year, month, report_date)
    ws['Z7'].font = Font(name='Calibri', size=7)
    ws['Z7'].alignment = Alignment(horizontal="center", vertical="center")
    wb.save(time_sheet_name)
