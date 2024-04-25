from calendar import monthrange
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Side

from db import add_employee_to_db, search_employee_in_db, search_last_name
from utils import check_employee, letter_code


# тонкая черная линия
stile_thin = Side(border_style="thin", color="000000")
# толстая черная линия
stile_medium = Side(border_style="medium", color="000000")


def load_info_employee(file_name, year, month, i):
    """Записывает в словарь дни отсутствия и причину из i-ой строки"""
    try:
        wb = load_workbook(file_name)
        ws = wb.active
    except:
        return {}
    # словарь для записи даты и причины отсутствия
    absence = {}
    # получает причину отсутствия
    if 'подразделение' in file_name:
        # получает причину отсутствия
        absence_reason = ws[f'E{i}'].value
        # дата начала отсутствия в виде строки
        string_start_absence = ws[f'H{i}'].value
        # дата окончания отсутствия в виде строки
        string_end_absence = ws[f'I{i}'].value
    else:
        absence_reason = ws[f'D{i}'].value
        string_start_absence = str(ws[f'G{i}'].value)
        string_end_absence = ws[f'H{i}'].value

    # преобразует ее в код
    absence_reason_code = letter_code(absence_reason)


    # преобразует дату начала отсутствия в формат datetime
    if type(string_start_absence) is datetime:
        date_start_absence = string_start_absence
    else:
        try:
            date_start_absence = datetime.strptime(string_start_absence, '%d.%m.%Y')
        except:
            date_start_absence = datetime.strptime(string_start_absence,'%Y-%m-%d %H:%M:%S')

    if type(string_end_absence) is datetime:
        date_end_absence = string_end_absence
    else:
        try:
            # преобразует дату окончания отсутствия в формат datetime
            if 'предположительно' in string_end_absence.lower():
                date_end_absence = datetime.strptime(
                    string_end_absence.lower(),
                    'предположительно до %d.%m.%Y 0:00:00'
                )
            else:
                date_end_absence = datetime.strptime(
                    string_end_absence.lower(), '%d.%m.%Y'
                )
        except:
            date_end_absence = datetime.strptime(string_end_absence,'%Y-%m-%d %H:%M:%S')

    # если дата начала отсутствия выпадает на текущий месяц,
    # то начинает счет с нее
    if datetime(year, month, monthrange(year, month)[1]) >= date_start_absence >= datetime(year, month, 1):
        current_date = date_start_absence
        # если дата конца больше или равна дате начала
        # и выпадает на текущий месяц
        if datetime(year, month, monthrange(year, month)[1]) >= date_end_absence >= current_date:
            # заканчиваем последней датой отстутствия
            end_date = date_end_absence
        # если дата конца отсутствия выпадает на следующий месяц
        elif date_end_absence > datetime(year, month, monthrange(year, month)[1]):
            # то последняя дата отсутствия - последний день текущего месяца
            end_date = datetime(year, month, monthrange(year, month)[1])
        else:
            return absence
    # если дата до начала текущего месяца
    elif date_start_absence < datetime(year, month, 1):
        # и дата конца после или равна началу текущего месяца
        if datetime(year, month, monthrange(year, month)[1]) >= date_end_absence >= datetime(year, month, 1):
            # то начинаем счет с первого дня месяца.
            current_date = datetime(year, month, 1)
            # заканчиваем последней датой отстутствия
            end_date = date_end_absence
        # если дата конца отсутствия выпадает на следующий месяц
        elif date_end_absence > datetime(year, month, monthrange(year, month)[1]):
            # то начинаем счет с первого дня месяца.
            current_date = datetime(year, month, 1)
            # то последняя дата отсутствия - последний день текущего месяца
            end_date = datetime(year, month, monthrange(year, month)[1])
        else:
            return absence
    else:
        return absence
    # проходимся по всем датам отсутствия
    while current_date != end_date + timedelta(days=1):
        # записывает дату и причину в словарь
        absence[current_date] = absence_reason_code
        # прибавляет один день
        current_date += timedelta(days=1)
    # возвращает словарь с датами отсутствия
    return absence


def search_all_employees(file_name, year, month):
    """Ищет всех сотрудников отдела в файле с отсутствующими"""
    employees = search_last_name()
    absence = {}
    i = 6
    try:
        if 'подразделение' in file_name:
            letter = 'A'
        else:
            letter = 'B'
        wb = load_workbook(file_name)
        ws = wb.active
    except:
        return {}
    while ws[f'{letter}{i}'].value is not None:
        string = ws[f'{letter}{i}'].value
        for employee in employees:
            if employee[6] == 'совм':
                if (employee[0] + ' ' + employee[1] + ' ' + employee[2] in string
                    and 'совм' in string):
                    key_name = f'{employee[0]} {employee[1][0]}.{employee[2][0]}. {employee[6]}'
                    absence[key_name] = {**absence.get(key_name, {}), **load_info_employee(file_name, year, month, i)}
                    break
            else: 
                if (not 'совм' in string and    
                    employee[0] + ' ' + employee[1] + ' ' + employee[2] in string):
                    key_name = f'{employee[0]} {employee[1][0]}.{employee[2][0]}.'
                    absence[key_name] = {**absence.get(key_name, {}), **load_info_employee(file_name, year, month, i)}
                    break
        i += 1
    return absence


def search_employee(absence: dict, name: str, internal_combine):
    """Ищет сотрудника в словаре отсутствующих сотрудников отдела
    и записывает все дни отсутствия с причинами в словарь"""
    if internal_combine == 'совм':
        key = f'{name} {internal_combine}'
    else:
        key = name
    if key in absence:
        return absence[key] 
    return {}   


# absence = search_all_employees("2024.04 Отсутствия сотрудников (1-15).xlsx", 2024, 4)
# for employee in absence.items():
#    print(employee)


# используется 
def add_data_from_db(file_name):
    """Добавляет сотрудников в базу из списка"""
    wb = load_workbook(file_name)
    ws = wb.active
    i = 2
    while ws[f'A{i}'].value is not None:
        name = ws[f'A{i}'].value
        last_name = name.split()[0]
        first_name = name.split()[1]
        patronymic = name.split()[2]
        time_sheet_number = ws[f'B{i}'].value
        position = ws[f'C{i}'].value
        part_time = ws[f'E{i}'].value
        internal_combine = ws[f'D{i}'].value
        if not internal_combine:
            internal_combine = '*'
        employment_date = ws[f'F{i}'].value
        dismissal_date = ws[f'G{i}'].value
        department = ws[f'H{i}'].value
        employee = search_employee_in_db(last_name, first_name, patronymic,
                                         time_sheet_number, position,
                                         part_time, internal_combine,
                                         department)
        if not check_employee(employee):
            add_employee_to_db(last_name, first_name, patronymic,
                               time_sheet_number, position, part_time,
                               internal_combine, employment_date,
                               dismissal_date, department)
        i += 1
 #   except Exception:
 #       print('Произошла ошибка добавления в базу данных')

